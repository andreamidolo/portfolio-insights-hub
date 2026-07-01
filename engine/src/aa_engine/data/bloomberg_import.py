"""ETL: Data Pack Bloomberg (Excel ``solo_valori``) → matrici ``date × ticker``.

Il file Excel esporta ogni strumento su **due righe**: una riga di DATE (con il
ticker nella colonna identificativa) e la riga successiva di VALORI. Ogni serie
porta le proprie date (start diversi per strumento), quindi l'allineamento è
esatto — niente ricostruzione del calendario.

Produce, sotto ``engine/data/market/`` (gitignored — dati Bloomberg licenziati):
    - ``prices.csv``        date × strumenti INVESTIBILI (base del motore)
    - ``signal_series.csv`` date × strumenti "solo_segnale" (indici/FX/commodity)
    - ``vol_indices.csv``   date × {VIX,VXN,MOVE,V2X,OVX,GVZ}
    - ``universe.csv``      manifest (ticker, nome, isin, valuta, asset_class, gruppo, ruolo)

Uso::

    python -m aa_engine.data.bloomberg_import /percorso/solo_valori.xlsx
"""

from __future__ import annotations

import argparse
import datetime as _dt
from pathlib import Path

import pandas as pd

# asset_class (foglio Universo) → gruppo di vincolo del motore (5 gruppi).
GROUP_OF_CLASS: dict[str, str] = {
    "Equity_DM": "equity",
    "Equity_EM": "equity",
    "Fixed_Income_IG": "fixed_income",
    "Fixed_Income_HY": "fixed_income",
    "Commodity": "commodities",
    "Alternative": "alternatives",
    # Signal_* → non investibili (regime/segnali), gruppo None.
}

VOL_TICKERS = {"VIX", "VXN", "MOVE", "V2X", "OVX", "GVZ"}


def _symbol(bbg_ticker: str) -> str:
    """Simbolo pulito = primo token del ticker Bloomberg ('SPY US Equity' → 'SPY')."""
    return str(bbg_ticker).strip().split()[0]


def _is_date(x) -> bool:
    return isinstance(x, (_dt.datetime, _dt.date))


def _is_num(x) -> bool:
    return isinstance(x, (int, float)) and not isinstance(x, bool)


def _parse_two_row_sheet(ws, ticker_col: int, series_start: int) -> dict[str, pd.Series]:
    """Estrae {symbol: Series(index=date, values=float)} da un foglio a coppie di righe.

    Cammina le righe: una riga con date nell'area-serie è una DATE-row (col
    ``ticker_col`` = ticker); la riga successiva con numeri è la sua VALUE-row.
    """
    out: dict[str, pd.Series] = {}
    pending_ticker: str | None = None
    pending_dates: list | None = None
    for row in ws.iter_rows(values_only=True):
        series = row[series_start:]
        first = next((v for v in series if v is not None), None)
        if _is_date(first):  # DATE-row
            pending_ticker = row[ticker_col]
            pending_dates = list(series)
            continue
        if _is_num(first) and pending_dates is not None:  # VALUE-row
            vals = list(series)
            pairs = [
                (d, float(v))
                for d, v in zip(pending_dates, vals)
                if _is_date(d) and _is_num(v)
            ]
            if pending_ticker and pairs:
                idx = pd.DatetimeIndex([p[0] for p in pairs])
                s = pd.Series([p[1] for p in pairs], index=idx).sort_index()
                s = s[~s.index.duplicated(keep="last")]
                out[_symbol(pending_ticker)] = s
            pending_ticker, pending_dates = None, None
    return out


def _detect_series_start(ws, ticker_col: int) -> int:
    """Trova la prima colonna della serie (dove la DATE-row ha una data)."""
    for row in ws.iter_rows(min_row=2, max_row=3, values_only=True):
        for j, v in enumerate(row):
            if _is_date(v):
                return j
    # fallback: subito dopo le colonne metadato
    return ticker_col + 4


def _to_frame(series_map: dict[str, pd.Series]) -> pd.DataFrame:
    if not series_map:
        return pd.DataFrame()
    df = pd.DataFrame(series_map).sort_index()
    df.index.name = "date"
    return df


def load_universe(ws) -> pd.DataFrame:
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        ac = str(r[0])
        rows.append({
            "symbol": _symbol(r[2]),
            "bbg_ticker": r[2],
            "name": r[3],
            "isin": r[4],
            "currency": r[5],
            "asset_class": ac,
            "sub_class": r[1],
            "role": r[6],
            "group": GROUP_OF_CLASS.get(ac),  # None per i Signal_*
        })
    return pd.DataFrame(rows)


def run(xlsx: str | Path, out_dir: str | Path) -> dict:
    import openpyxl

    xlsx = Path(xlsx)
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)

    universe = load_universe(wb["Universo_Completo"])
    universe.to_csv(out_dir / "universe.csv", index=False)

    # A_Prezzi: ticker in col idx1, serie da ~idx5
    a = wb["A_Prezzi"]
    a_start = _detect_series_start(a, ticker_col=1)
    prices_all = _parse_two_row_sheet(a, ticker_col=1, series_start=a_start)

    # separa investibili vs solo_segnale via manifest
    invest = set(universe.loc[universe["role"] == "investibile", "symbol"])
    signal = set(universe.loc[universe["role"] == "solo_segnale", "symbol"])
    prices = _to_frame({k: v for k, v in prices_all.items() if k in invest})
    sig_from_a = {k: v for k, v in prices_all.items() if k in signal}

    # B_Indici_Vol: ticker in col idx0, serie da ~idx4
    b = wb["B_Indici_Vol"]
    b_start = _detect_series_start(b, ticker_col=0)
    vol_all = _parse_two_row_sheet(b, ticker_col=0, series_start=b_start)
    vol = _to_frame({k: v for k, v in vol_all.items() if k in VOL_TICKERS})
    sig_from_b = {k: v for k, v in vol_all.items() if k not in VOL_TICKERS}

    signals = _to_frame({**sig_from_a, **sig_from_b})

    prices.to_csv(out_dir / "prices.csv")
    vol.to_csv(out_dir / "vol_indices.csv")
    if not signals.empty:
        signals.to_csv(out_dir / "signal_series.csv")

    def _cov(df):
        if df.empty:
            return {}
        return {"n": df.shape[1], "rows": df.shape[0],
                "start": str(df.index[0].date()), "end": str(df.index[-1].date())}

    return {
        "prices": _cov(prices), "vol_indices": _cov(vol), "signals": _cov(signals),
        "universe_rows": len(universe),
        "investable_found": sorted(prices.columns.tolist()),
        "investable_missing": sorted(invest - set(prices.columns)),
        "out_dir": str(out_dir),
    }


def _main() -> None:
    from rich.console import Console

    p = argparse.ArgumentParser(description="ETL Data Pack Bloomberg → CSV date×ticker.")
    p.add_argument("xlsx", help="percorso del file Excel (solo_valori)")
    p.add_argument("--out", default=None, help="cartella output (default engine/data/market)")
    args = p.parse_args()
    out = args.out or (Path(__file__).resolve().parents[3] / "data" / "market")
    info = run(args.xlsx, out)
    c = Console()
    c.print(info)


if __name__ == "__main__":
    _main()
